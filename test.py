from openpyxl import Workbook
import openpyxl

# # 创建一个新的Excel工作簿
# workbook = Workbook()
# sheet = workbook.active

# # 写入表头
# headers = ["Role ID", "Name", "Image", "Height", "Speed X", "Speed Y", "Skills", "Boss"]
# sheet.append(headers)

# # 遍历roles字典并将数据写入Excel
# roles = {
#     "role_0": {
#         "name": "召唤",
#         "image": "static/npc/npc0.png",
#         "height": 142,
#         "speed_x": 680,
#         "speed_y": 840,
#         "skills": "s",
#         "boss": "s",
#     },
#     "role_1": {
#         "name": "气功",
#         "image": "static/npc/npc1.png",
#         "height": 158,
#         "speed_x": 700,
#         "speed_y": 835,
#         "skills": "z",
#         "boss": "w",
#     },
#     "role_2": {
#         "name": "战法",
#         "image": "static/npc/npc2.png",
#         "height": 142,
#         "speed_x": 700,
#         "speed_y": 790,
#         "skills": "weh",
#         "boss": "f",
#     },
#     "role_3": {
#         "name": "奶萝",
#         "image": "static/npc/npc3.png",
#         "height": 142,
#         "speed_x": 720,
#         "speed_y": 875,
#         "skills": "w",
#         "boss": "s",
#     },
#     "role_4": {
#         "name": "剑魔",
#         "image": "static/npc/npc4.png",
#         "height": 168,
#         "speed_x": 680,
#         "speed_y": 830,
#         "skills": "dy",
#         "boss": "r",
#     },
#     "role_5": {
#         "name": "气功2",
#         "image": "static/npc/npc5.png",
#         "height": 158,
#         "speed_x": 650,
#         "speed_y": 850,
#         "skills": "sqx",
#         "boss": "w",
#     },
#     "role_6": {
#         "name": "四姨",
#         "image": "static/npc/npc6.png",
#         "height": 170,
#         "speed_x": 635,
#         "speed_y": 780,
#         "skills": "whyx",
#         "boss": "e",
#     },
#     "role_7": {
#         "name": "刃影",
#         "image": "static/npc/npc7.png",
#         "height": 164,
#         "speed_x": 635,
#         "speed_y": 780,
#         "skills": "th",
#         "boss": "w",
#     },
#     "role_8": {
#         "name": "女漫游",
#         "image": "static/npc/npc8.png",
#         "height": 178,
#         "speed_x": 630,
#         "speed_y": 770,
#         "skills": "x",
#         "boss": "e",
#     },
#     "role_9": {
#         "name": "红眼",
#         "image": "static/npc/npc9.png",
#         "height": 180,
#         "speed_x": 630,
#         "speed_y": 780,
#         "skills": "sh",
#         "boss": "e",
#     },
#     "role_10": {
#         "name": "审判",
#         "image": "static/npc/npc10.png",
#         "height": 180,
#         "speed_x": 650,
#         "speed_y": 760,
#         "skills": "t",
#         "boss": "d",
#     },
#     "role_11": {
#         "name": "奶妈",
#         "image": "static/npc/npc11.png",
#         "height": 173,
#         "speed_x": 740,
#         "speed_y": 870,
#         "skills": "r",
#         "boss": "d",
#     },
# }

# for role_id, role_info in roles.items():
#     row_data = [
#         role_id,
#         role_info["name"],
#         role_info["image"],
#         role_info["height"],
#         role_info["speed_x"],
#         role_info["speed_y"],
#         role_info["skills"],
#         role_info["boss"],
#     ]
#     sheet.append(row_data)

# # 保存Excel文件
# workbook.save("roles_data.xlsx")


# 读取角色信息的函数
def load_roles_from_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook["roles"]

    roles = {}

    for row in sheet.iter_rows(
        min_row=2, values_only=True
    ):  # 从第二行开始读取数据，第一行是标题
        role_id = row[0]  # 假设第一列是角色ID（如 "role_0"）
        role_info = {
            "name": row[1],
            "image": "static/roles/" + row[0],
            "height": row[2],
            "speed_x": row[3],
            "speed_y": row[4],
            "skills": row[5],
            "boss": row[6],
        }
        roles[role_id] = role_info

    return roles


# 在加载角色前先调用该函数
roles = load_roles_from_excel("roles_data.xlsx")
print(roles)
